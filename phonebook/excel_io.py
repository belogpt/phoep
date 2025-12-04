"""Импорт и экспорт телефонной книги в формате Excel."""
import io
import os
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook

from .models import Contact
from .repository import save_contacts

COLUMNS = [
    'Department',
    'Name',
    'Office Number',
    'Mobile Number',
    'Other Number',
    'Head Portrait',
]


def export_to_excel(contacts: List[Contact]) -> bytes:
    """Возвращает байты Excel-файла со всеми контактами."""
    data = []
    for contact in contacts:
        data.append({
            'Department': contact.group,
            'Name': contact.name,
            'Office Number': contact.office,
            'Mobile Number': contact.mobile,
            'Other Number': contact.other,
            'Head Portrait': '',
        })
    df = pd.DataFrame(data, columns=COLUMNS)
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    return buffer.read()


def import_from_excel(file_stream) -> int:
    """Импортирует контакты из Excel, полностью перезаписывая XML. Возвращает количество контактов."""
    df = pd.read_excel(file_stream, engine=None)
    # нормализуем имена колонок
    columns_map = {c.lower(): c for c in df.columns}
    required = {
        'department': 'Department',
        'name': 'Name',
        'office number': 'Office Number',
        'mobile number': 'Mobile Number',
        'other number': 'Other Number',
        'head portrait': 'Head Portrait',
    }
    for key in required:
        if key not in columns_map:
            raise ValueError('Неверный формат столбцов Excel')
    contacts: List[Contact] = []
    for _, row in df.iterrows():
        department = str(row[columns_map['department']]).strip() if not pd.isna(row[columns_map['department']]) else ''
        name = str(row[columns_map['name']]).strip() if not pd.isna(row[columns_map['name']]) else ''
        office = '' if pd.isna(row[columns_map['office number']]) else str(row[columns_map['office number']]).strip()
        mobile = '' if pd.isna(row[columns_map['mobile number']]) else str(row[columns_map['mobile number']]).strip()
        other = '' if pd.isna(row[columns_map['other number']]) else str(row[columns_map['other number']]).strip()
        if not department:
            continue
        if not any(num.strip() for num in [office, mobile, other]):
            continue
        contacts.append(Contact(group=department, name=name, office=office, mobile=mobile, other=other, photo=''))
    save_contacts(contacts)
    return len(contacts)


@dataclass
class RawContact:
    """Контакт, полученный из «сырой» таблицы отделов."""

    full_department_name: str
    full_name: str
    raw_row_data: Dict[str, str] = field(default_factory=dict)
    internal_extension: Optional[str] = None


def _is_department_cell(cell) -> bool:
    """Определяет, что ячейка B относится к строке с названием отдела."""

    value = cell.value
    if not isinstance(value, str) or not value.strip():
        return False

    fill = getattr(cell, "fill", None)
    if fill and getattr(fill, "patternType", None) == "solid":
        fg_color = getattr(fill, "fgColor", None)
        if fg_color and getattr(fg_color, "type", None) in {"theme", "rgb"}:
            return True

    if not re.search(r"\d", value):
        return True

    return False


def extract_internal_extension_from_row(
    phone_internal_raw: str, row_values: List[str]
) -> Optional[str]:
    """Извлекает внутренний номер (3–5 цифр) из строки сотрудника."""

    def _cleanup_digits(text: str) -> str:
        return re.sub(r"\D", "", text)

    if phone_internal_raw:
        digits = _cleanup_digits(str(phone_internal_raw))
        if digits.isdigit() and 3 <= len(digits) <= 5:
            return digits

    for value in row_values:
        if not value:
            continue
        for match in re.findall(r"[\d\s\+\-\(\)]+", value):
            digits = _cleanup_digits(match)
            if digits.isdigit() and 3 <= len(digits) <= 5:
                return digits

    return None


def parse_raw_department_table(filepath: str) -> List[RawContact]:
    """Парсит «сырую» Excel-таблицу с отделами и сотрудниками."""

    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)

    wb = load_workbook(filepath)
    ws = wb.active
    current_department_full_name: Optional[str] = None
    raw_contacts: List[RawContact] = []

    for row_idx in range(8, ws.max_row + 1):
        if ws.cell(row=row_idx, column=4).value == "Справочно:":
            break

        cell_b = ws.cell(row=row_idx, column=2)
        val_b = cell_b.value

        if _is_department_cell(cell_b):
            current_department_full_name = str(val_b).strip()
            continue

        if isinstance(val_b, (int, float)) and current_department_full_name:
            full_name = str(ws.cell(row=row_idx, column=4).value or "").strip()
            position = str(ws.cell(row=row_idx, column=5).value or "").strip()
            phone_external = str(ws.cell(row=row_idx, column=6).value or "").strip()
            phone_internal_raw = str(ws.cell(row=row_idx, column=7).value or "").strip()
            email = str(ws.cell(row=row_idx, column=8).value or "").strip()

            if not full_name:
                continue

            row_values = [
                str(ws.cell(row=row_idx, column=col).value or "")
                for col in range(4, 9)
            ]
            internal_extension = extract_internal_extension_from_row(
                phone_internal_raw, row_values
            )

            raw_contacts.append(
                RawContact(
                    full_department_name=current_department_full_name,
                    full_name=full_name,
                    raw_row_data={
                        "position": position,
                        "phone_external": phone_external,
                        "phone_internal_raw": phone_internal_raw,
                        "email": email,
                        "row_index": row_idx,
                    },
                    internal_extension=internal_extension,
                )
            )

    if not raw_contacts:
        raise ValueError(
            "Не удалось распознать ни одного контакта. Проверьте формат файла."
        )

    return raw_contacts


def normalize_raw_contacts(raw_contacts: List[RawContact], dept_alias_map: Dict[str, str]) -> List[Contact]:
    """Преобразует RawContact в модель Contact с использованием сокращений отделов."""

    grouped: Dict[str, List[Contact]] = {}
    group_order: List[str] = []

    for raw in raw_contacts:
        group = dept_alias_map.get(raw.full_department_name, raw.full_department_name)
        contact = Contact(
            group=group,
            name=raw.full_name,
            office=raw.internal_extension or "",
            mobile="",
            other="",
            photo="",
        )
        if not any(num.strip() for num in [contact.office, contact.mobile, contact.other]):
            continue

        if group not in grouped:
            grouped[group] = []
            group_order.append(group)
        grouped[group].append(contact)

    ordered: List[Contact] = []
    for group in group_order:
        contacts = grouped[group]
        contacts.sort(key=lambda c: c.name.casefold())
        ordered.extend(contacts)

    return ordered
